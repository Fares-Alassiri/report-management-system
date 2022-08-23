import os
# import pandas as pd
import openpyxl

from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.contrib.auth.forms import PasswordChangeForm

from django.contrib import messages

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required

from django.contrib.auth.models import Group as UserGroup
from core.models import *
from core.forms import CreateUserForm, EditProfileEmail

from app import settings

# Create your views here.

@login_required(login_url='signin')
def home(request):
    return render(request, 'core/dashboard.html')

@login_required(login_url='signin')
def reports(request):
    profile = Profile.objects.get(user=request.user)
    user_groups = Group.objects.filter(profiles=profile)
    reports = Report.objects.filter(groups__in=user_groups).order_by('-created_at')
    context = {
        'reports': reports,
    }
    return render(request, 'core/reports.html', context)

@login_required(login_url='signin')
def report(request, pk):
    report = get_object_or_404(Report, id=pk)
    report_groups = Group.objects.filter(report=report).all()
    flag = False
    for group in report_groups:
        obj = Profile.objects.filter(profiles=group).all()
        if Profile.objects.get(user=request.user) in obj:
            flag = True
    if flag == False:
        messages.success(request, 'You have not permission to view the report!')
        return redirect('reports')
    context = {
        'report': report,
    }
    return render(request, 'core/report.html', context)

@login_required(login_url='signin')
def create_report(request):
    if request.method == 'POST':
        try:
            request.POST.get('Name')[0]
            with transaction.atomic():
                report = Report.objects.create(
                    name = request.POST.get('Name'),
                    author = Profile.objects.get(user=request.user),
                    content = request.POST.get('Content'),
                )

                tags = str(request.POST.get('Tags')).split(',')
                for t in tags:
                    tag, created_tag = Tag.objects.get_or_create(name=t)
                    report.tags.add(tag)
                
                groups = request.POST.getlist('Groups')
                for g in groups:
                    group = Group.objects.get(id=int(g))
                    report.groups.add(group)

                report.editors.add(Profile.objects.get(user=request.user))

                files = request.FILES.getlist('Files')
                for file in files:
                    Attachment.objects.create(
                        file = file,
                        report = report
                    )
        except:
            groups = Group.objects.filter(profiles=Profile.objects.get(user=request.user))
            context = { 
                'groups': groups,
                'error': 'Please make sure that you entered valid inputs and try again'
            }

            return render(request, 'core/create_report.html', context)
        messages.success(request, 'Report created successfully')
        return redirect('/report')
    groups = Group.objects.filter(profiles=Profile.objects.get(user=request.user))
    context = { 
        'groups': groups,
    }

    return render(request, 'core/create_report.html', context)

@login_required(login_url='signin')
def update_report(request, pk):
    if request.method == 'POST':
        try:
            request.POST.get('Name')[0]
            with transaction.atomic():
                report = Report.objects.get(id=pk)

                report.name = request.POST.get('Name')
                report.content = request.POST.get('Content')

                report.tags.clear()
                tags = str(request.POST.get('Tags')).split(',')
                for t in tags:
                    tag, created_tag = Tag.objects.get_or_create(name=t)
                    report.tags.add(tag)
                
                report.groups.clear()
                groups = request.POST.getlist('Group')
                groups[0]
                for g in groups:
                    group = Group.objects.get(id=int(g))
                    report.groups.add(group)

                report.editors.add(Profile.objects.get(user=1))
                report.editors.add(Profile.objects.get(user=1))

                report.save()

                files = request.FILES.getlist('Files')
                for file in files:
                    Attachment.objects.create(
                        file = file,
                        report = report
                    )
        except Exception as e:
            report = get_object_or_404(Report, id=pk)
            report_groups = Group.objects.filter(report=report).all()
            flag = False
            for group in report_groups:
                obj = Profile.objects.filter(profiles=group).all()
                if Profile.objects.get(user=request.user) in obj:
                    flag = True
            if flag == False:
                messages.success(request, 'You have not permission to edit the report!')
                return redirect('reports')
            groups = Group.objects.filter(profiles=Profile.objects.get(user=request.user))
            rgroups = report.groups.all()
            temp3 = []
            for element in rgroups:
                if element not in groups:
                    temp3.append(element)
            dgroups = temp3
            context = { 
                'report': report,
                'groups': groups,
                'rgroups': rgroups,
                'dgroups': dgroups,
                'error': 'Please make sure that you entered valid inputs and try again'
            }

            return render(request, 'core/update_report.html', context)
        # Attachment.objects.create(file=request.FILES.get('Files'))
        messages.success(request, 'The report has been updated successfully')
        return redirect('/report/{}'.format(pk))
    report = get_object_or_404(Report, id=pk)
    report_groups = Group.objects.filter(report=report).all()
    flag = False
    for group in report_groups:
        obj = Profile.objects.filter(profiles=group).all()
        if Profile.objects.get(user=request.user) in obj:
            flag = True
    if flag == False:
        messages.success(request, 'You have not permission to edit the report!')
        return redirect('reports')
    groups = Group.objects.filter(profiles=Profile.objects.get(user=request.user))
    rgroups = report.groups.all()
    temp3 = []
    for element in rgroups:
        if element not in groups:
            temp3.append(element)
    dgroups = temp3
    context = { 
        'report': report,
        'groups': groups,
        'rgroups': rgroups,
        'dgroups': dgroups
    }

    return render(request, 'core/update_report.html', context)

@login_required(login_url='signin')
def delete_report(request, pk):
    report = get_object_or_404(Report, id=pk)
    report_groups = Group.objects.filter(report=report).all()
    flag = False
    for group in report_groups:
        obj = Profile.objects.filter(profiles=group).all()
        if Profile.objects.get(user=request.user) in obj:
            flag = True
    if flag == False:
        messages.success(request, 'You have not permission to delete the report!')
        return redirect('reports')
    report.delete()
    messages.success(request, 'Report deleted successfully')
    return redirect('/report')

@login_required(login_url='signin')
def delete_attachment(request, rid, aid):
    attachment = get_object_or_404(Attachment, id=aid)
    report = get_object_or_404(Report, id=rid)
    report_groups = Group.objects.filter(report=report).all()
    flag = False
    for group in report_groups:
        obj = Profile.objects.filter(profiles=group).all()
        if Profile.objects.get(user=request.user) in obj:
            flag = True
    if flag == False:
        messages.success(request, 'You have not permission to edit the report!')
        return redirect('reports')
    attachment.delete()
    messages.success(request, 'Attachment deleted successfully')
    return redirect('/report/{}/update'.format(rid))

def signin(request):
    if request.user.is_authenticated:
        return redirect('reports')
    else:
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')
            
            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                return redirect('reports')
            else:
                context = {'error': 'Invalid username or password'}
                return render(request, 'core/login.html', context= context)

        context = {}
        return render(request, 'core/login.html', context= context)

@login_required(login_url='signin')
def signout(request):
    logout(request)
    return redirect('signin')

def signup(request):
    if request.user.is_authenticated:
        return redirect('reports')
    else:
        form = CreateUserForm()

        if request.method == 'POST':
            try:
                with transaction.atomic():
                    form = CreateUserForm(request.POST)
                    if form.is_valid():
                        form.save()
                        user = form.cleaned_data.get('username')
                        Profile.objects.create(user= User.objects.get(username=user))
                        user_group = UserGroup.objects.get(name='User') 
                        user_group.user_set.add(User.objects.get(username=user))
                        messages.success(request, 'Account was created successfully for '+user)
                        return redirect('signin')
            except:
                context = {'form': form, 'error': 'Please make sure that you entered valid inputs and try again'}
                return render(request, 'core/signup.html', context= context)
            

        context = {'form': form}
        return render(request, 'core/signup.html', context= context)

def account(request):
    emailform = EditProfileEmail(instance= request.user)
    passform = PasswordChangeForm(user= request.user)
    context = {'passform': passform, 'emailform':emailform,}
    return render(request, 'core/account.html', context)

def editemail(request):
    if request.method == 'POST':
        form = EditProfileEmail(request.POST, instance= request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'The email was successfully updated')
            return redirect('account')
        messages.success(request, 'The email was not updated, try again')
        return redirect('account')

def editpass(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'The password was successfully updated')
            return redirect('signin')
        messages.error(request, 'The password was not updated, try again and make sure you entered correct old password and re write new one correct')
        return redirect('account')
    
def accounts_login(request):
    return redirect('signin')

@login_required
def reports_new(request):
    ugroup = None
    if request.user.groups.exists():
        ugroup = request.user.groups.all()[0].name
    if ugroup != 'Admin':
        messages.success(request, 'You have not permission to go to previous page')
        return redirect('reports')
    if request.method == 'POST':
        try:
            with transaction.atomic():
                file1 = request.FILES.getlist('zip')
                file2 = request.FILES.get('excel')
                if not file2:
                    return render(request, 'core/new_reports.html', {'error': 'Please upload at least the excel file'})
                if not (allowed_excel_file(str(file2))):
                    return render(request, 'core/new_reports.html', {'error': 'Please upload first file with correct extensions .xlsx'})
                xlsx = openpyxl.load_workbook(file2)
                sheet = xlsx.active
                idxs = []
                for row in sheet.iter_rows():
                    for cell in row:
                        idxs.append(cell.value)
                    break

                if idxs[0] != 'ReportName' or idxs[1] != 'ReportContent' or idxs[2] != 'ReportTags' or idxs[3] != 'ReportGroups' or idxs[4] != 'Attachments_id':
                    return render(request, 'core/new_reports.html', {'error': 'Please upload first file with first row same as picture'})
                i = 0
                for row in sheet.iter_rows():
                    if i == 0:
                        i += 1
                        continue
                    vals = []
                    for cell in row:
                        vals.append(cell.value)
                    report = Report.objects.create(
                        name = vals[0],
                        author = Profile.objects.get(user=request.user),
                        content = vals[1],
                        attachment_ref = vals[4]
                    )

                    tags = str(vals[2]).split(',')
                    for t in tags:
                        tag, created_tag = Tag.objects.get_or_create(name=t)
                        report.tags.add(tag)
                    
                    groups = str(vals[3]).split(',')
                    for g in groups:
                        group, group_tag = Group.objects.get_or_create(name=g)
                        report.groups.add(group)

                    report.editors.add(Profile.objects.get(user=request.user))
                
                if file1:
                    for file in file1:
                        report_attachment_id = str(file).split('.')[0].split('_')[0]
                        report = Report.objects.filter(attachment_ref=report_attachment_id).order_by('-created_at').first()
                        Attachment.objects.create(
                            file = file,
                            report = report
                        )
                messages.success(request, 'The new reports were created successfully')
                return redirect('reports')
        except Exception as e:
            return render(request, 'core/new_reports.html', {'error':'There is error, try again and make sure you follow the instructions'})    
    return render(request, 'core/new_reports.html', {})

def allowed_zip_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in {'zip'}

def allowed_excel_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in {'xlsx'}